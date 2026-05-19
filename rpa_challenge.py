"""
rpa_challenge.py — AgentFarm RPA Challenge Automation
Navigates to https://rpachallenge.com/, downloads the challenge Excel,
reads the 10 employee records, and fills all 10 form rounds with Playwright.

Audit outputs (screenshot + recording) are uploaded to Azure Blob Storage.
Container: ${AZURE_AUDIT_CONTAINER}  (default: agent-audit-logs)
Blob path:  rpa-sessions/{YYYY-MM-DD}/{session_id}/
"""
import asyncio
import io
import os
import sys
import uuid
from datetime import datetime, timezone

import openpyxl
import requests
from playwright.async_api import async_playwright

EXCEL_URL = "https://rpachallenge.com/assets/downloadFiles/challenge.xlsx"
CHALLENGE_URL = "https://rpachallenge.com/"

# Map Excel column headers → Angular ng-reflect-name attribute values
FIELD_MAP = {
    "First Name":       "labelFirstName",
    "Last Name":        "labelLastName",
    "Company Name":     "labelCompanyName",
    "Role in Company":  "labelRole",
    "Address":          "labelAddress",
    "Email":            "labelEmail",
    "Phone Number":     "labelPhone",
}


def _blob_client():
    """Return a ContainerClient, or None if blob storage is not configured."""
    conn_str = os.environ.get("AZURE_STORAGE_CONNECTION_STRING", "")
    if not conn_str or "YOUR_ACCOUNT" in conn_str:
        return None
    try:
        from azure.storage.blob import ContainerClient
        container = os.environ.get("AZURE_AUDIT_CONTAINER", "agent-audit-logs")
        client = ContainerClient.from_connection_string(conn_str, container)
        client.create_container()          # no-op if already exists
        return client
    except Exception as exc:
        print(f"[blob] WARNING: could not connect to blob storage: {exc}")
        return None


def upload_bytes(client, blob_path: str, data: bytes, content_type: str) -> str | None:
    """Upload bytes to blob, return public URL or None on failure."""
    if client is None:
        return None
    try:
        client.upload_blob(blob_path, data, overwrite=True,
                           content_settings={"content_type": content_type})
        url = f"{client.url}/{blob_path}"
        print(f"[blob] Uploaded → {url}")
        return url
    except Exception as exc:
        print(f"[blob] WARNING: upload failed for {blob_path}: {exc}")
        return None


def upload_file(client, blob_path: str, local_path: str, content_type: str) -> str | None:
    """Upload a local file to blob, delete it locally on success."""
    if client is None or not os.path.exists(local_path):
        return None
    with open(local_path, "rb") as f:
        url = upload_bytes(client, blob_path, f.read(), content_type)
    if url:
        os.remove(local_path)
        print(f"[blob] Local file removed: {local_path}")
    return url


def download_excel() -> list[dict]:
    """Download challenge.xlsx and return list of row dicts."""
    print("[rpa] Downloading Excel from rpachallenge.com...")
    resp = requests.get(EXCEL_URL, timeout=30)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in ws[1]]
    people: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            people.append(dict(zip(headers, row)))

    print(f"[rpa] Loaded {len(people)} records:")
    for p in people:
        print(f"      {p.get('First Name')} {p.get('Last Name')} | {p.get('Company Name')}")
    return people


async def run_challenge(people: list[dict], session_id: str) -> None:
    date_prefix = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    blob_prefix = f"rpa-sessions/{date_prefix}/{session_id}"
    client = _blob_client()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=False,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--window-size=1280,700"],
        )
        page = await browser.new_page(viewport={"width": 1280, "height": 700})

        print(f"[rpa] Opening {CHALLENGE_URL}")
        await page.goto(CHALLENGE_URL, wait_until="networkidle", timeout=30000)
        await page.wait_for_timeout(1500)

        await page.click("button:has-text('Start')")
        print("[rpa] Clicked Start — challenge begun")
        await page.wait_for_timeout(1200)

        for i, person in enumerate(people, 1):
            fname = person.get("First Name", "")
            lname = person.get("Last Name", "")
            print(f"[rpa] Round {i}/{len(people)}: {fname} {lname}")

            for col_name, ng_name in FIELD_MAP.items():
                value = person.get(col_name)
                if value is None:
                    value = ""
                selector = f"input[ng-reflect-name='{ng_name}']"
                await page.fill(selector, str(value))

            await page.click("input[value='Submit']")
            await page.wait_for_timeout(800)

        print("[rpa] All rounds complete — checking result...")
        await page.wait_for_timeout(3000)

        # Screenshot → bytes → blob (no temp file written to disk)
        screenshot_bytes = await page.screenshot()
        blob_path = f"{blob_prefix}/final-score.png"
        url = upload_bytes(client, blob_path, screenshot_bytes, "image/png")
        if url:
            print(f"[rpa] Screenshot uploaded: {url}")
        else:
            # Fallback: save locally if blob not configured
            local_path = "/tmp/rpa_final_score.png"
            await page.screenshot(path=local_path)
            print(f"[rpa] Screenshot saved locally: {local_path}")

        await browser.close()

    # Upload recording if it exists (written by ffmpeg in run_rpa_demo.sh)
    recording_path = "/tmp/rpa_recording.mp4"
    if os.path.exists(recording_path):
        blob_path = f"{blob_prefix}/recording.mp4"
        url = upload_file(client, blob_path, recording_path, "video/mp4")
        if url:
            print(f"[rpa] Recording uploaded: {url}")
        else:
            print(f"[rpa] Recording kept locally: {recording_path}")


if __name__ == "__main__":
    people = download_excel()
    if not people:
        print("[rpa] ERROR: No data read from Excel", file=sys.stderr)
        sys.exit(1)
    session_id = os.environ.get("RPA_SESSION_ID", uuid.uuid4().hex[:12])
    print(f"[rpa] Session ID: {session_id}")
    asyncio.run(run_challenge(people, session_id))
    print("[rpa] Done.")

